import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog
from ttkthemes import ThemedStyle
from tkinter.scrolledtext import ScrolledText
import openpyxl as xl
import sendSMS
import sendMMS



#-------Helper class and methods-------#

#holds all information needed to send messages
class submission:
    def __init__(self):
        #login info
        self.account_sid = ""
        self.auth_token = ""
        self.fromNum = ""

        #excel data
        self.xlsx = None
        self.numbers = []

        #message info
        self.message = ""
        self.urls = [""]

        #output info
        self.numsNotSent = []
        self.messageCount = 0
        self.returnMessage = ""

        #check if running
        self.running = False

    #resets important information
    def reset(self):
        self.numsNotSent = []
        self.numbers = []
        self.messageCount = 0
        self.running = False

#login function for account information
def login(sub, self):
    sub.account_sid = self.sidEntry.get()
    sub.auth_token = self.tokenEntry.get()
    sub.fromNum = ("+1" + self.fromNumEntry.get())
    self.controller.show_frame("subPage")
    return()

#pulls numbers from excel file and send message
def send(sub, self):

    #disable button during process
    self.sumbitButton.configure(state="disabled")
    sub.running = True

    #get message and URL
    sub.message = self.messageBox.get("1.0", "end-1c")
    sub.urls[0] = self.urlField.get()

    #load excel file
    try:
        wrkbk = xl.load_workbook(sub.xlsx.name)
        wb = wrkbk.active 

        #get each number
        for i in range(1, wb.max_row+1):
            cell_obj = wb.cell(row=i, column=3)
            sub.numbers.append("+1" + str(cell_obj.value))

    #if excel file is not correct
    except:
        self.updateOutput("messages not sent, check xlsx file info\n")
        self.sumbitButton.configure(state="!disabled")
        return()
    
    #set size of progress bar
    self.progress.configure(maximum=wb.max_row)
    
    #check if message exists
    if(len(sub.message) <= 0):
        self.updateOutput("messages not sent, no message provided\n")
        self.sumbitButton.configure(state="!disabled")
        sub.reset()
        return()

    #call send
    if(sub.urls[0] != ""):
        #send mms if there is a URL image
        self.updateOutput(sendMMS.sendMMS(self, sub))
    else:
        self.updateOutput(sendSMS.sendSMS(self, sub))

    #enable button and reset
    self.sumbitButton.configure(state="!disabled")
    sub.reset()
    return()
    
#brings up file explorer for an excel file
def browseXLSX(sub, self):
    #set file object based on file input
    sub.xlsx = filedialog.askopenfile(  initialdir = "/",
                                        title = "Select a xlsx File",
                                        filetypes = [("xlsx files", "*.xlsx*")])
      
    # Change label contents
    if(sub.xlsx != None):
        self.xlsxButton.configure(text=getFileName(sub.xlsx.name))
    else:
        self.xlsxButton.configure(text="Browse xlsx File")

    return()   

#gets name of file without full path
def getFileName(name):
    index = len(name)-1
    retName = ""
    while(name[index] != "/"):
        retName += name[index]
        index -= 1
    return(retName[::-1])

#halts sending process
def cancel(sub, self):
    if(sub.running):
        self.updateOutput("task halted\n")
        sub.running = False



#-------GUI setup-------#

#mainApp class that holds all frames
class mainApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.sub = submission()

        self.title("Auto Messenger")
        self.geometry("800x650")
        self.style = ThemedStyle()
        self.style.theme_use("radiance")

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        #get all frames for swapping
        self.frames = {}
        for F in (loginPage, subPage):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("loginPage")

    #show new frame
    def show_frame(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()


#login Frame
class loginPage(ttk.Frame):
    def __init__(self, parent, controller):
        ttk.Frame.__init__(self, parent)
        self.controller = controller

        self.sidLabel = ttk.Label(self, text="Enter Account SID: ", font=(20))
        self.sidLabel.grid(column=0, row=0, padx = 20, pady=10, sticky="e")
        
        self.sidEntry = ttk.Entry(self, width=50)
        self.sidEntry.grid(column=1, row=0, padx = 20, pady=10, sticky="w" )

        self.tokenLabel = ttk.Label(self, text="Enter Auth Token: ", font=(20))
        self.tokenLabel.grid(column=0, row=1, padx = 20, pady=10, sticky="e")

        self.tokenEntry = ttk.Entry(self, width=50)
        self.tokenEntry.grid(column=1, row=1, padx = 20, pady=10, sticky="w")

        self.fromNumLabel = ttk.Label(self, text="Enter 10 Digit Number to Text From: ", font=(20))
        self.fromNumLabel.grid(column=0, row=2, padx = 20, pady=10, sticky="e")

        self.fromNumEntry = ttk.Entry(self, width=50)
        self.fromNumEntry.grid(column=1, row=2, padx = 20, pady=10, sticky="w")

        self.loginButton = ttk.Button(self, text="Login", command = lambda:login(controller.sub, self))
        self.loginButton.grid(column=1, row=3, padx = 20, pady=10, sticky="w")

        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=5)

#submission frame
class subPage(ttk.Frame):
    def __init__(self, parent, controller):
        ttk.Frame.__init__(self, parent)
        self.controller = controller

        self.title = ttk.Label(self, text="Enter Message Below: ", font=(20))
        self.title.grid(column=0, row=0, padx=10, pady=5, sticky="w")

        self.messageBox = ScrolledText(self, width=50,  height=20)
        self.messageBox.grid(column=0, row=1, padx=10, pady=5)

        self.xlsxButton = ttk.Button(self, text="Browse xlsx File", command = lambda:browseXLSX(controller.sub, self))
        self.xlsxButton.grid(column=1, row=1, padx=10, pady=25, sticky="n")

        self.urlLabel = ttk.Label(self, text="Image URL:")
        self.urlLabel.grid(column=1, row=1,padx=10, pady=100, sticky="n"+"w")

        self.urlField = ttk.Entry(self, width=50)
        self.urlField.grid(column=1, row=1, padx=10, pady=125, sticky="n"+"w")

        self.sumbitButton = ttk.Button(self, text='Send Message', command = lambda:send(controller.sub, self))
        self.sumbitButton.grid(column=1, row=1, padx=10, pady=25, sticky="s")

        self.cancelButton = ttk.Button(self, text='Stop', command = lambda:cancel(controller.sub, self))
        self.cancelButton.grid(column=1, row=2, padx=10, pady=10)

        self.ResultLabel = ttk.Label(self, text="Results: ", font=(20))
        self.ResultLabel.grid(column=0, row=2, padx=10, sticky="w")

        self.progress = ttk.Progressbar(self, orient='horizontal', mode='determinate', length=400)
        self.progress.grid(column=0, row=3, padx=10, sticky="w")

        self.outputLabel = ttk.Label(self, text='\n')
        self.outputLabel.grid(column=0, row=4, padx=10, pady=10, sticky="w")

        empty = ttk.Label(self, text='\n\n\n')
        empty.grid(column=0, row=5)

        self.backButton = ttk.Button(self, text='Return to Login', command = lambda:controller.show_frame("loginPage"))
        self.backButton.grid(column=1, row=6, padx=10, pady=10)

        self.hintLabel = tk.Message(self, text="HINT: \nxlsx file should contain 10 digit numbers in the 3rd column" +\
                                                 "\nif number of messages not sent is high, check formating", \
                                       font=("", 8), width=400)
        self.hintLabel.grid(column=0, row=6, padx=10, pady=10, sticky="sw")


    #updates the output label text
    def updateOutput(self, message):
        self.outputLabel.configure(text=message)
        root.update()


#main function driver
if __name__ == "__main__":
    root = mainApp()
    root.mainloop()